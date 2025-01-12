from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime as dt
from openpyxl import load_workbook
from time import sleep
import pandas as pd
import schedule
import os

# Funções do Webdriver
def start_driver():
    chrome_options = Options()
    arguments = ['--lang=pt-BR', '--start-maximized']

    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option('prefs', {
        'download.directory_upgrade': True,
        'download.prompt_for_download': False,
        'profile.default_content_setting_values.notifications': 2,
        'profile.default_content_setting_values.automatic_downloads': 1,
    })

    driver = webdriver.Chrome(options=chrome_options)

    return driver


def wait_for_element(driver, by, value, timeout=60):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )


def formatted_price(price):
    price_before = price.replace('R$', '').strip()
    price_before = price_before.replace(',', '.')
    return float(price_before)


# Função principal
def main():
    link = 'https://www.terabyteshop.com.br/produto/31959/gabinete-gamer-redragon-wideload-extreme-mid-tower-rgb-vidro-curvado-temperado-atx-black-sem-fonte-sem-fan-ca-605b'
    driver = start_driver()
    driver.get(link)
    sleep(0.5)
    price_in_cash = driver.execute_script(f'''
return document.evaluate(
    '//p[@id= "valVista"]', 
    document, 
    null, 
    XPathResult.FIRST_ORDERED_NODE_TYPE, 
    null
    ).singleNodeValue;
                                    ''')
    driver.execute_script('arguments[0].scrollIntoView({behavior: "instant", block: "center"});', price_in_cash)
    price_in_cash = driver.execute_script('return arguments[0].innerText;', price_in_cash)
    price_full = driver.find_element(By.XPATH, '(//span[@id="valParc"])[2]').text

    # Formatando os preços
    price_in_cash = formatted_price(price_in_cash)
    price_full = formatted_price(price_full)

    print(f'Preço à vista: {price_in_cash};\nPreço cheio: {price_full}')
    date = dt.now().strftime('%d/%m/%Y %H:%M')

    # Caminho do arquivo Excel
    path = os.getcwd() + '/price_gabinete.xlsx'

    try:
        # Se o arquivo já existir
        wb = load_workbook(path)
        sheet = wb.active

        # Verificar o último ID e incrementar
        last_row = sheet.max_row
        new_id = sheet.cell(row=last_row, column=1).value + 1

        # Adicionando a nova linha
        new_line = [new_id, date, price_in_cash, price_full, link]
        sheet.append(new_line)
        wb.save(path)
        print('Linha adicionada com sucesso!')
    except FileNotFoundError:
        # Criando o arquivo Excel se não existir
        new_id = 1
        df = pd.DataFrame([{
            'id': new_id,
            'Date': date,
            'Price in cash': price_in_cash,
            'Price full': price_full,
            'Link': link
        }])
        df.to_excel(path, index=False)
        print('Arquivo Excel criado com sucesso!')


schedule.every(1).day.do(main)

print(f'Próxima verificação irá ocorrer às: {schedule.next_run()}')

main()

while True:
    schedule.run_pending()
    sleep(1)